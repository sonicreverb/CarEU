import os.path
from _init_ import BASE_DIR
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import load_workbook


def create_driver():
    chrome_options = webdriver.ChromeOptions()
    prefs = {"profile.managed_default_content_settings.images": 2}
    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    return driver


def get_htmlsoup(url, driver):
    try:
        driver.get(url)
        time.sleep(1)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        driver.close()
        return soup

    except:
        print('ERROR: invalid url', url)
        return 'error invalid url'


def get_data(url, driver):
    soup = get_htmlsoup(url, driver)
    if soup == 'error invalid url':
        return None

    title = soup.find('h1', class_='h2 g-col-8').get_text()
    brutto_price = ''.join(soup.find('span', class_='brutto-price u-text-bold').get_text()[:-10].split())
    netto_price = ''.join(soup.find('p', class_='netto-price').get_text()[:-9].split())

    soup_img_li = soup.find_all('div', class_='gallery-bg js-gallery-img js-load-on-demand')
    img_li = []
    for img_index in range(len(soup_img_li) // 2 - 1):
        img_li.append(soup_img_li[img_index].get('data-src'))

    technical_options_soup_li = soup.find('div', class_='vip-details-block u-margin-bottom-18') \
        .find_all('span', {'class': 'g-col-6'})
    TO_dict = {}
    # меня напрягает и фрустрирует, что линтер ругается на это, поэтому ЭТО здесь
    col_key = 'error'

    for col_index in range(len(technical_options_soup_li)):
        if col_index % 2 == 0:
            col_key = technical_options_soup_li[col_index].get_text()
        else:
            col_value = technical_options_soup_li[col_index].get_text().replace(u'\u2009', u' ')
            TO_dict[col_key] = col_value.replace(u'\xa0', u' ')

    characteristics = ''
    for row in soup.find_all('p', class_='bullet-point-text'):
        characteristics += row.get_text() + '; '

    # todo !
    description_soup = soup.find('div', class_='description-text js-original-description g-col-12')
    description_headlines = description_soup.find_all('b')
    description_ul = description_soup.find_all('ul')
    description_str = ''

    if len(description_ul) == 0:
        description_str = soup.find('div', class_='description-text js-original-description g-col-12').get_text()
    else:
        for index in range(len(description_headlines)):
            # todo !!
            if index >= len(description_ul):
                break
            description_str += description_headlines[index].get_text() + '\n'
            for li in description_ul[index].find_all('li'):
                description_str += li.get_text() + '\n'
            description_str += '\n'

    dealer_name = soup.find('div', class_='vip-box seller-box cBox cBox--content hidden-s u-clearfix'). \
        find('a', {'data-google-interstitial': 'false'}).get_text()

    dealer_phone = ''.join(soup.find('p', class_='seller-phone u-margin-bottom-9 u-text-orange u-text-bold').get_text()
                           .split())
    dealer_link = soup.find('div', class_='vip-box seller-box cBox cBox--content hidden-s u-clearfix') \
        .find('a').get('href')

    product_data = {'Title': title, 'URL': url, 'BruttoPrice': brutto_price, 'NettoPrice': netto_price, 'ImgLi': img_li,
                    'TechOptDict': TO_dict, 'CharacteristicsStr': characteristics,
                    'DescriptionStr': description_str.strip(), 'DealerData': [dealer_name, dealer_link, dealer_phone]}

    return product_data


# def write_product_to_excel(product_data):
#     XLSX_PATH = os.path.join(BASE_DIR, 'scraper', 'excel tables', 'product_data.xlsx')
#     workbook = load_workbook(XLSX_PATH)
#     worksheet = workbook['data']
#
#     line_num = str(worksheet.max_row + 1)
#     print(line_num)
#
#     worksheet['A' + line_num] = product_data['Title']
#     # todo!
#     worksheet['B' + line_num] = product_data['Title'].split()[0]
#     worksheet['C' + line_num] = product_data['Title'].split()[1]
#     worksheet['D' + line_num] = product_data['URL']
#     worksheet['E' + line_num] = product_data['BruttoPrice']
#     worksheet['F' + line_num] = product_data['NettoPrice']
#     worksheet['G' + line_num] = product_data['CharacteristicsStr']
#     worksheet['H' + line_num] = product_data['DescriptionStr']
#
#     img_li = product_data['ImgLi']
#     for j in range(15):
#         if j >= len(img_li):
#             break
#         # столбцы D-R
#         worksheet[chr(73 + j) + str(line_num)] = img_li[j]
#
#     worksheet['X' + line_num] = product_data['DealerData'][1]
#     worksheet['Y' + line_num] = product_data['DealerData'][2]
#     worksheet['Z' + line_num] = str(datetime.now())[:-7]
#     worksheet['BA' + line_num] = "Да"
#
#     TO_keys = []
#
#     # чтение значений с таблицы от АА1 до конца имеющихся значений (не больше АZ1)
#     COLUM_NUM_TO_CHAR_CONST = 65
#     for col_index in range(26):
#         cell_value = str(worksheet['A' + chr(COLUM_NUM_TO_CHAR_CONST + col_index) + '1'].value)
#         if cell_value == "None":
#             break
#         TO_keys.append(cell_value)
#
#     TO_dict = product_data['TechOptDict']
#     print(TO_keys)
#     print(TO_dict)
#
#     for key in TO_dict:
#         value = TO_dict[key]
#         write_key_to_table_flag = False
#
#         # проверка, что ключ надо записать в таблицу и значения в диапазоне до AZ1
#         if key not in TO_keys and len(TO_keys) < 26:
#             TO_keys.append(key)
#             write_key_to_table_flag = True
#
#         if key in TO_keys:
#             index = TO_keys.index(key)
#             worksheet["A" + chr(index + COLUM_NUM_TO_CHAR_CONST) + line_num] = value
#             if write_key_to_table_flag:
#                 worksheet["A" + chr(index + COLUM_NUM_TO_CHAR_CONST) + '1'] = key
#
#     workbook.save(XLSX_PATH)
#     workbook.close()
#
#
# def execute_scraper():
#     line = 0
#     XLSX_PATH = os.path.join(BASE_DIR, 'scraper', 'excel tables', 'product_data.xlsx')
#     CSV_PATH = os.path.join(BASE_DIR, 'scraper', 'excel tables', 'product_data.csv')
#     with open(os.path.join(BASE_DIR, "scraper", "links", "active_links.txt")) as al_input:
#         for product_link in al_input:
#             try:
#                 product_data = get_data(product_link.strip())
#                 write_product_to_excel(product_data)
#                 if line % 100 == 0:
#                     upload_csv_to_sheets(XLSX_PATH, CSV_PATH)
#                 print(product_link)
#                 line += 1
#             except:
#                 print("error, could't get data", product_data)
#         upload_csv_to_sheets(XLSX_PATH, CSV_PATH)
#
#
# execute_scraper()
