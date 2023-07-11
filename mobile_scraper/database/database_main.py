import datetime
import os.path
import time

import psycopg2
import re
import openpyxl
import os.path as osph

from openpyxl.utils import get_column_letter
from mobile_scraper.database.config import host, user, password, db_name
from pycbrf import ExchangeRates

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))


# получение соеднения с БД
def get_connection_to_db():
    try:
        # соединение с существующей базой данных
        connection = psycopg2.connect(
            host=host,
            user=user,
            password=password,
            database=db_name
        )

        connection.autocommit = True
        print("\n[PostGreSQL INFO] Connected successfully.")
        return connection
    except Exception as _ex:
        print("[PostGreSQL INFO] Error while working with PostgreSQL", _ex)
        return None


# создание/обновление таблицы моделей в БД
def update_models_table_to_db(models_data):
    # получаем соединение
    connection = get_connection_to_db()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # проверка на существование таблицы
            table_name = "vehicles_models"
            cursor.execute("SELECT EXISTS(SELECT relname FROM pg_class WHERE relname=%s)", (table_name,))
            table_exists = cursor.fetchone()[0]

            if table_exists:
                # если таблица существует - удаляем все данные перед записью
                cursor.execute(f"DELETE FROM {table_name};")
            else:
                # иначе создаём новую таблицу
                cursor.execute(f"CREATE TABLE {table_name}(id serial PRIMARY KEY, producer varchar(255) NOT NULL,"
                               f"model varchar(255) NOT NULL, model_id int, url varchar(511));")

            # запись в БД
            for tmp in range(len(models_data["Producer"])):
                producer = models_data['Producer'][tmp]
                model = models_data['Models'][tmp]
                model_id = models_data['ModelID'][tmp]
                url = models_data['URL'][tmp]

                # SQL запрос
                query = "INSERT INTO vehicles_models (producer, model, model_id, url) VALUES (%s, %s, %s, %s)"
                print(query, (producer, model, model_id, url))
                cursor.execute(query, (producer, model, model_id, url))

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")


# возвращает результат выполняемого запроса в БД
def get_querry_result(querry):
    # получаем соединение
    connection = get_connection_to_db()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            result = []

            # выполнение БД запроса
            cursor.execute(querry)

            try:
                for bd_object in cursor.fetchall():
                    result.append(bytes(bd_object[0], 'utf-8').decode('unicode_escape'))

                print("[PostGreSQL INFO] Data was read successfully.")
            except psycopg2.ProgrammingError:
                pass

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
        return result
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")
        return None


# получение словаря с моделями из БД
def read_models_from_db():
    # получение марок из БД
    producers_li = get_querry_result("SELECT producer from vehicles_models;")

    # получение моделей из БД
    models_li = get_querry_result("SELECT model from vehicles_models;")

    if producers_li and models_li:
        # инициализация словаря формата - модель: [марка_модели1, марка_модели2, ...]
        producers_models_data = {producer: [] for producer in set(producers_li)}

        # заполнение словаря моделями
        for indx in range(len(producers_li)):
            producers_models_data[producers_li[indx]].append(models_li[indx])

        return producers_models_data
    else:
        print('Error while trying to read models or producer in read_models_from_db().')
        return None


# получение цены расстоможки согласно объёму из БД
def get_custom_clearance_coeff(volume):
    # получаем соединение
    connection = get_connection_to_db()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # проверка на существование таблицы
            table_name = "tcalc"
            cursor.execute("SELECT EXISTS(SELECT relname FROM pg_class WHERE relname=%s)", (table_name,))
            table_exists = cursor.fetchone()[0]

            if table_exists:
                # получение цены расстоможки согласно объёму из БД

                cursor.execute(
                    f"SELECT sumrub FROM tcalc WHERE volume = {volume};"
                )

                coeff = float(cursor.fetchone()[0])

                print("[PostGreSQL INFO] Data was read successfully.")

            else:
                print(f"[PostGreSQL INFO] Error while trying to read {table_name}. {table_name} doesn't exist.")
                connection.close()
                return None

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")

        return coeff
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")
        return None


# чтение локальных ссылок из БД
def get_local_links_from_db():
    try:
        return get_querry_result("SELECT source_url from vehicles_data;")
    except Exception as _ex:
        print('Error in get_local_links from db', _ex)
        return None


# возвращает массив активных ссылок из БД
def get_active_links_from_db():
    try:
        return get_querry_result("SELECT source_url from vehicles_data WHERE activity = true;")
    except Exception as _ex:
        print('Error in get_active_links from db', _ex)
        return None


# возвращает массив неактивных ссылок из БД
def get_unactive_links_from_db():
    try:
        return get_querry_result("SELECT source_url from vehicles_data WHERE activity = false AND activity_validation "
                                 "= false;")
    except Exception as _ex:
        print('Error in get_active_links from db', _ex)
        return None


# чтение активных наименований из БД
def get_active_names_from_db():
    try:
        return get_querry_result("SELECT name FROM vehicles_data WHERE activity = true;")
    except Exception as _ex:
        print('Error in get_active_links from db', _ex)
        return None


# запись данных машины в БД
def write_productdata_to_db(product_data):
    # получаем соединение
    connection = get_connection_to_db()

    # запись в БД
    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:

            # ОСНОВНАЯ ИНФОРМАЦИЯ
            name = product_data['Title']
            url = product_data['URL']
            brutto_price = product_data['BruttoPrice']
            netto_price = product_data['NettoPrice']

            # ПРОИЗВОДИТЕЛЬ И МОДЕЛЬ
            all_models_dict = read_models_from_db()
            product_make = ' '
            product_model = ' '

            # todo убедиться в правильности
            for make in all_models_dict:
                if make in name.split():
                    product_make = make
                    for model in all_models_dict[make]:
                        if model in name:
                            product_model = model

            # ПОБОЧКА
            characteristics = product_data['CharacteristicsStr']
            description = product_data['DescriptionStr']
            description_trd = product_data['DescriptionTRD']

            dealer_name = product_data['DealerData'][0]
            dealer_url = product_data['DealerData'][1]
            dealer_phone = product_data['DealerData'][2]

            # ИЗОБРАЖЕНИЯ
            img1 = img2 = img3 = img4 = img5 = img6 = img7 = img8 = img9 = img10 = img11 = img12 = img13 = img14 \
                = img15 = ''

            if len(product_data['ImgLi']) >= 1:
                img1 = product_data['ImgLi'][0]
            if len(product_data['ImgLi']) >= 2:
                img2 = product_data['ImgLi'][1]
            if len(product_data['ImgLi']) >= 3:
                img3 = product_data['ImgLi'][2]
            if len(product_data['ImgLi']) >= 4:
                img4 = product_data['ImgLi'][3]
            if len(product_data['ImgLi']) >= 5:
                img5 = product_data['ImgLi'][4]
            if len(product_data['ImgLi']) >= 6:
                img6 = product_data['ImgLi'][5]
            if len(product_data['ImgLi']) >= 7:
                img7 = product_data['ImgLi'][6]
            if len(product_data['ImgLi']) >= 8:
                img8 = product_data['ImgLi'][7]
            if len(product_data['ImgLi']) >= 9:
                img9 = product_data['ImgLi'][8]
            if len(product_data['ImgLi']) >= 10:
                img10 = product_data['ImgLi'][9]
            if len(product_data['ImgLi']) >= 11:
                img11 = product_data['ImgLi'][10]
            if len(product_data['ImgLi']) >= 12:
                img12 = product_data['ImgLi'][11]
            if len(product_data['ImgLi']) >= 13:
                img13 = product_data['ImgLi'][12]
            if len(product_data['ImgLi']) >= 14:
                img14 = product_data['ImgLi'][13]
            if len(product_data['ImgLi']) >= 15:
                img15 = product_data['ImgLi'][14]

            # КАТЕГОРИИ
            techopt = product_data['TechOptDict']
            category1 = category2 = category3 = category4 = category5 = category6 = category7 = category8 = \
                category9 = category10 = category11 = category12 = category13 = category14 = category15 = category16 \
                = category17 = category18 = category19 = category20 = category21 = category22 = category23 = ''

            if 'Категория' in techopt:
                category1 = techopt['Категория']
            if 'Первая регистрация' in techopt:
                category2 = techopt['Первая регистрация']
            if 'Коробка передач' in techopt:
                category3 = techopt['Коробка передач']
            if 'Топливо' in techopt:
                category4 = techopt['Топливо']
            if 'Пробег' in techopt:
                category5 = techopt['Пробег']
            if 'Мощность' in techopt:
                category6 = techopt['Мощность']
            if 'Объем двигателя' in techopt:
                category7 = techopt['Объем двигателя']
                # получаем объём двигателя для дальнейших калькуляций
                volume_str = techopt['Объем двигателя'].replace(" ", "")
                match = re.search(r'\d+', volume_str)
                if match:
                    volume = int(match.group())
            if 'Количество мест' in techopt:
                category8 = techopt['Количество мест']
            if 'Число дверей' in techopt:
                category9 = techopt['Число дверей']
            if 'Класс экологической безопасности' in techopt:
                category10 = techopt['Класс экологической безопасности']
            if 'Количество владельцев транспортного средства' in techopt:
                category11 = techopt['Количество владельцев транспортного средства']
            if 'Общий осмотр' in techopt:
                category12 = techopt['Общий осмотр']
            if 'Цвет' in techopt:
                category13 = techopt['Цвет']
            if 'Цвет по классификации производителя' in techopt:
                category14 = techopt['Цвет по классификации производителя']
            if 'Состояние транспортного средства' in techopt:
                category15 = techopt['Состояние транспортного средства']
            if 'Оригинал' in techopt:
                category16 = techopt['Оригинал']
            if 'Потребление' in techopt:
                category17 = techopt['Потребление']
            if 'CO₂ Emissions' in techopt:
                category18 = techopt['CO₂ Emissions']
            if 'Датчики парковки' in techopt:
                category19 = techopt['Датчики парковки']
            if 'Дизайн салона' in techopt:
                category20 = techopt['Дизайн салона']
            if 'Номер транспортного средства' in techopt:
                category21 = techopt['Номер транспортного средства']
            if 'Наличие' in techopt:
                category22 = techopt['Наличие']
            if 'Вид основного топлива' in techopt:
                category23 = techopt['Вид основного топлива']

            # UPLOAD PRICES DATA
            euro_rate = get_euro_rate()

            brutto_price_rubles = (int(product_data['BruttoPrice']) * euro_rate)

            # Цена с комиссией =((нетто + брутто*0,1)+(брутто*0,07+300))*курсевро
            comission_price_rubles = ((int(product_data["NettoPrice"]) + int(product_data['BruttoPrice']) * 0.1) +
                                      (int(product_data['BruttoPrice']) * 0.07 + 300)) * euro_rate

            # Цена с доставкой = цена с комиссией + 2000евро
            delivery_price_rubles = int(comission_price_rubles + 2000 * euro_rate)

            # Цена с растоможкой = цена с доставкой + стоимость растоможки согласно объему двигателя + 50000р
            custom_clearance_coeff = get_custom_clearance_coeff(volume)
            if custom_clearance_coeff:
                custom_clearance_rubles = delivery_price_rubles + custom_clearance_coeff + 50000
            else:
                custom_clearance_rubles = 0

            # SQL запрос
            query = "INSERT INTO vehicles_data (name, make, model, source_url, price_brutto, price_netto, " \
                    "characteristics, description, dealer_name, dealer_url, dealer_phone, img1, img2, img3, img4, " \
                    "img5, img6, img7, img8, img9, img10, img11, img12, img13, img14, img15, category1, category2," \
                    " category3, category4, category5, category6, category7, category8, category9, category10," \
                    " category11, category12, category13, category14, category15, category16, category17," \
                    " category18, category19, category20, category21, category22, category23, description_translated," \
                    " price_brutto_rubles, commission_price_rubles, delivery_price_rubles, customs_clearance_price)" \
                    " VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s," \
                    "%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, " \
                    "%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

            cursor.execute(query, (name, product_make, product_model, url, brutto_price, netto_price, characteristics,
                                   description, dealer_name, dealer_url, dealer_phone, img1, img2, img3, img4, img5,
                                   img6, img7, img8, img9, img10, img11, img12, img13, img14, img15, category1,
                                   category2, category3, category4, category5, category6, category7, category8,
                                   category9, category10, category11, category12, category13, category14, category15,
                                   category16, category17, category18, category19, category20, category21, category22,
                                   category23, description_trd, int(brutto_price_rubles), int(comission_price_rubles),
                                   int(delivery_price_rubles), int(custom_clearance_rubles)))
            print("[PostGreSQL INFO] Data was wrote to DataBase.")
        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")


# отмечает товар неактивым в БД по его сслыке
def edit_product_activity_in_db(local_link, activity_status=False, unactivity_valid=False):
    # получаем соединение
    connection = get_connection_to_db()
    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # проверка на существование таблицы
            table_name = "vehicles_data"
            cursor.execute("SELECT EXISTS(SELECT relname FROM pg_class WHERE relname=%s)", (table_name,))
            table_exists = cursor.fetchone()[0]
            if table_exists:
                if activity_status:
                    # отметка товара активным в БД
                    cursor.execute(
                        "UPDATE vehicles_data SET activity = true WHERE source_url = "
                        f"'{local_link}';"
                    )

                if not activity_status:
                    # отметка товара неактивным в БД
                    cursor.execute(
                        "UPDATE vehicles_data SET activity = false WHERE source_url = "
                        f"'{local_link}';"
                    )
                    # установка даты, с которой товар стал неактивен в БД
                    cursor.execute(
                        "UPDATE vehicles_data SET unactive_since = NOW() WHERE activity = false AND unactive_since is "
                        "NULL;"
                    )

                if unactivity_valid:
                    cursor.execute(
                        "UPDATE vehicles_data SET activity_validation = true WHERE source_url = "
                        f"'{local_link}';"
                    )
            else:
                print(f"[PostGreSQL INFO] Error while trying to read {table_name}. {table_name} doesn't exist.")
                connection.close()
                return None

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")
        return None


# удаляет неактивные позиции в БД по истечении суток
def delete_unactive_positions():
    # получаем соединение
    connection = get_connection_to_db()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # выполнение БД запроса
            cursor.execute("DELETE FROM vehicles_data WHERE unactive_since <= NOW() - INTERVAL '1 day' AND "
                           "activity_validation = true;")

            print("[PostGreSQL INFO] Data was deleted successfully.")

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")
        return None


# получение текущего курса
def get_euro_rate():
    # response = requests.get("https://www.cbr-xml-daily.ru/daily_json.js").json()
    # eur_rate = response["Valute"]["EUR"]["Value"]
    # return float(eur_rate)
    date = datetime.datetime.now().strftime("%Y-%m-%d")
    rates = ExchangeRates(date)
    currency_data = list(filter(lambda el: el.code == "EUR", rates.rates))[0]
    return float(currency_data.value)


# создание таблицы таможенного калькулятора
def create_tcalc():
    # получаем соединение
    connection = get_connection_to_db()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # проверка на существование таблицы
            table_name = "tcalc"
            cursor.execute("SELECT EXISTS(SELECT relname FROM pg_class WHERE relname=%s)", (table_name,))
            table_exists = cursor.fetchone()[0]

            if table_exists:
                # если таблица существует - удаляем все данные перед записью
                cursor.execute(f"DELETE FROM {table_name};")

                # получаем текущий курс евро
                euro_rate = get_euro_rate()

                # запись в БД с кодировкой
                for vol in range(500, 8200 + 1):
                    if vol < 1000:
                        rate = 1.5
                    elif vol < 1501:
                        rate = 1.7
                    elif vol < 1801:
                        rate = 2.5
                    elif vol < 2301:
                        rate = 2.7
                    elif vol < 3001:
                        rate = 3
                    else:
                        rate = 3.6

                    # SQL запрос
                    query = "INSERT INTO tcalc (volume, rate, sum, sumRUB) VALUES (%s, %s, %s, %s)"
                    # print(query, (vol, rate, vol*rate, vol*rate*euro_rate + 17200))
                    cursor.execute(query, (vol, rate, vol * rate, vol * rate * euro_rate + 17200))
            else:
                # иначе создаём новую таблицу
                # cursor.execute(f"CREATE TABLE {table_name}(id serial PRIMARY KEY, producer varchar(255) NOT NULL,"
                #                f"model varchar(255) NOT NULL, model_id int, url varchar(511));")
                print("[PostGreSQL INFO] Error, table doesn't exist...")

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")


# обновление таможенного калькулятора согласно текущему курсу
def update_tcalc():
    while True:
        # получаем соединение
        connection = get_connection_to_db()

        # если соединение установлено успешно
        if connection:
            with connection.cursor() as cursor:
                # получаем текущий курс евро
                euro_rate = get_euro_rate()

                for vol in range(500, 8200 + 1):
                    if vol < 1000:
                        rate = 1.5
                    elif vol < 1501:
                        rate = 1.7
                    elif vol < 1801:
                        rate = 2.5
                    elif vol < 2301:
                        rate = 2.7
                    elif vol < 3001:
                        rate = 3
                    else:
                        rate = 3.6

                    # SQL запрос
                    query = f"UPDATE tcalc SET sumrub = {vol * rate * euro_rate + 17200} WHERE volume = {vol};"
                    # print(query, (vol, rate, vol*rate, vol*rate*euro_rate + 17200))
                    cursor.execute(query)

            connection.commit()
            # прикрываем соединение
            connection.close()
            print('[PostGreSQL INFO] tcalc UPDATE COMPLETE.')
            print("[PostGreSQL INFO] Connection closed.")
            print(f"\n[TCALC UPDATER] Таможенный калькулятор успешно обновлен: {datetime.datetime.now()}")
            time.sleep(30 * 60)
        else:
            print("[PostGreSQL INFO] Error, couldn't get connection...")
            print(f"\n[TCALC UPDATER] ERROR! Не удалось обновить таможенный калькулятор.")


# обновление цен в таблице
def update_final_prices():
    # получаем соединение
    connection = get_connection_to_db()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # получаем текущий курс евро
            euro_rate = get_euro_rate()

            # SQL на чтение из таблицы имён запрос
            query = f"SELECT id, price_brutto, price_netto, category7 FROM vehicles_data;"
            cursor.execute(query)
            # перерасчёт цен с учётом нового курса евро
            products_data = cursor.fetchall()
            for product in products_data:
                product_id = product[0]
                brutto = product[1]
                netto = product[2]
                volume_str = product[3].replace(" ", "")
                match = re.search(r'\d+', volume_str)
                if match:
                    volume = int(match.group())

                brutto_price_rub = (int(brutto) * euro_rate)

                # Цена с комиссией =((нетто + брутто*0,1)+(брутто*0,07+300))*курсевро
                comission_price_rubles = ((int(netto) + int(brutto) * 0.1) +
                                          (int(brutto) * 0.07 + 300)) * euro_rate

                # Цена с доставкой = цена с комиссией + 2000евро
                delivery_price_rubles = int(comission_price_rubles + 2000 * euro_rate)

                # Цена с растоможкой = цена с доставкой + стоимость растоможки согласно объему двигателя + 50000р
                custom_clearance_coeff = get_custom_clearance_coeff(volume)
                if custom_clearance_coeff:
                    custom_clearance_rubles = delivery_price_rubles + custom_clearance_coeff + 50000
                else:
                    custom_clearance_rubles = 0

                # SQL запрос на обновление цен
                query = f"UPDATE vehicles_data SET price_brutto_rubles = {brutto_price_rub}, commission_price_rubles" \
                        f" = {comission_price_rubles}, delivery_price_rubles = {delivery_price_rubles}, " \
                        f"customs_clearance_price = {custom_clearance_rubles} WHERE id = '{product_id}';"

                cursor.execute(query)

        connection.commit()
        # прикрываем соединение
        connection.close()
        print('[PostGreSQL INFO] vehicles_data UPDATE COMPLETE.')
        print("[PostGreSQL INFO] Connection closed.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")


# запись данных из БД в xlsx файл querry - запрос в БД, filename - название файла в который запись
def write_data_to_xlsx(querry, filename):
    # получаем соединение
    connection = get_connection_to_db()

    euro_rate = get_euro_rate()

    # если соединение установлено успешно
    if connection:
        with connection.cursor() as cursor:
            # SQL запрос
            cursor.execute(querry)
            rows = cursor.fetchall()

            # создание XLSX-файла
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # запись заголовков столбцов
            columns = [i[0] for i in cursor.description]
            for column_index, column_name in enumerate(columns, start=1):
                column_letter = get_column_letter(column_index)
                worksheet[f"{column_letter}1"] = column_name
                worksheet["BF1"] = 'euro_rate'

            # запись данных
            for row_index, row in enumerate(rows, start=2):
                for column_index, cell_value in enumerate(row):
                    column_letter = get_column_letter(column_index + 1)
                    try:
                        worksheet[f"{column_letter}{row_index}"] = str(cell_value)
                    except Exception as _ex:
                        print('[XLSX FILE] could\'t write cell value, string error.')
                    worksheet[f"BF{row_index}"] = str(euro_rate)

            # сохранение файла
            filename_path = osph.join(BASE_DIR, 'mobile_scraper', 'database', filename)
            workbook.save(filename_path)

        connection.commit()
        # прикрываем соединение
        connection.close()
        print("[PostGreSQL INFO] Connection closed.")
        print("[XLSX FILE] XLSX file was updated successfully.")
    else:
        print("[PostGreSQL INFO] Error, couldn't get connection...")


# записывает все данные из БД в 5 файлов формата xlsx
def upload_db_data_to_xlsx():
    makes_category1 = ['Audi', 'Renault', 'Saab']
    querry1 = "SELECT * FROM vehicles_data WHERE " + ' OR '.join(
        ["make = '{}'".format(make) for make in makes_category1]) + ';'
    write_data_to_xlsx(querry1, 'output1.xlsx')

    makes_category2 = ['Mercedes-Benz', 'Mazda', 'MINI', 'Mitsubishi']
    querry2 = "SELECT * FROM vehicles_data WHERE " + ' OR '.join(
        ["make = '{}'".format(make) for make in makes_category2]) + ';'
    write_data_to_xlsx(querry2, 'output2.xlsx')

    makes_category3 = ['Skoda', 'Opel', 'Nissan', 'Peugeot', 'Porsche', 'Smart', 'Subaru', 'Suzuki', 'Isuzu',
                       'Lamborghini',
                       'KTM', 'Cobra', 'Corvette', 'Ferrari', 'Koenigsegg', 'Lotus', 'Maserati', 'Maybach', 'McLaren',
                       'Plymouth', 'Pontiac', 'Rolls-Royce', 'Ruf', 'Bugatti']
    querry3 = "SELECT * FROM vehicles_data WHERE " + ' OR '.join(
        ["make = '{}'".format(make) for make in makes_category3]) + ';'
    write_data_to_xlsx(querry3, 'output3.xlsx')

    makes_category4 = ['BMW', 'Volvo', 'Toyota', 'ALPINA', 'Alfa Romeo', 'Aston Martin', 'Bentley', 'Cadillac',
                       'Chevrolet',
                       'Chrysler', 'Citroën', 'Dodge', 'Fiat', 'Infiniti', 'Honda', 'Hyundai']
    querry4 = "SELECT * FROM vehicles_data WHERE " + ' OR '.join(
        ["make = '{}'".format(make) for make in makes_category4]) + ';'
    write_data_to_xlsx(querry4, 'output4.xlsx')

    makes_category5 = ['Land Rover', 'Lexus', 'Lincoln', 'Jaguar', 'Jeep', 'Kia', 'Volkswagen', 'Ford']
    querry5 = "SELECT * FROM vehicles_data WHERE " + ' OR '.join(
        ["make = '{}'".format(make) for make in makes_category5]) + ';'
    write_data_to_xlsx(querry5, 'output5.xlsx')
